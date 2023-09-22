import pandas as pd
import numpy as np
from datetime import datetime, timedelta,date
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st                
import boto3
from io import BytesIO
from tempfile import NamedTemporaryFile
import time
import  streamlit_tree_select
import copy
import streamlit.components.v1 as components
from calendar import monthrange
s3 = boto3.client('s3')

#---------------------------define parameters--------------------------
st.set_page_config(
    page_title="Sabra HealthCare monthly reporting App",
    page_icon="🧊",
    initial_sidebar_state="expanded",
    layout="wide")
placeholder = st.empty()

st.title("Sabra HealthCare Monthly Reporting App")
sheet_name_account_mapping="Account_Mapping"
sheet_name_entity_mapping="Property_Mapping"
sheet_name_BPC_pull="BPC_pull"
sheet_name_format='Format'
sheet_name_discrepancy="Discrepancy_Review"
bucket_mapping="sabramapping"
bucket_PL="operatorpl"

@st.cache_data
def get_operator_list(bucket_mapping):
    operatorlist = s3.get_object(Bucket=bucket_mapping, Key="Initial_info.xlsx")
    operator_list = pd.read_excel(operatorlist['Body'].read(), sheet_name='Operator_List')
    return operator_list
operator_list=get_operator_list(bucket_mapping)

col1,col2=st.columns(2)
with col1:
    operator= st.selectbox('Operator Name',(operator_list))

@st.cache_data
def Initial_Paramaters(operator):
    # drop down list of operator
    if operator!='select operator':
        mapping_path="Mapping/"+operator+"/"+operator+"_Mapping.xlsx"
        PL_path=operator+"/"+operator+"_P&L.xlsx"
        Discrepancy_path="Total_Diecrepancy_Review.xlsx"
        BPCpull =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
        BPC_pull=pd.read_excel(BPCpull['Body'].read(),sheet_name=sheet_name_BPC_pull,header=0)
        BPC_pull=BPC_pull.set_index(["ENTITY","ACCOUNT"])
        BPC_pull.columns=list(map(lambda x :str(x),BPC_pull.columns))
        
        # read format table
        mapping_format =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
        format_table=pd.read_excel(mapping_format['Body'].read(), sheet_name=sheet_name_format,header=0)
        
        month_dic={10:["october","oct","10/","-10","/10","10"],11:["november","nov","11/","-11","/11","11"],12:["december","dec","12/","-12","/12","12"],1:["january","jan","01/","1/","-1","-01","/1","/01"],\
                   2:["february","feb","02/","2/","-2","-02","/2","/02"],3:["march","mar","03/","3/","-3","-03","/3","/03"],4:["april","apr","04/","4/","-4","-04","/4","/04"],\
                   5:["may","05/","5/","-5","-05","/5","/05"],6:["june","jun","06/","6/","-06","-6","/6","/06"],\
                   7:["july","jul","07/","7/","-7","-07","/7","/07"],8:["august","aug","08/","8/","-8","-08","/8","/08"],9:["september","sep","09/","9/","-09","-9","/9","/09"]}
        year_dic={2021:["2021","21"],2022:["2022","22"],2023:["2023","23"],2024:["2024","24"],2025:["2025","25"],2026:["2026","26"],2019:["2019","19"],2018:["2018","18"],2020:["2020","20"]} 

    else:
        st.stop()
    return PL_path,Discrepancy_path,mapping_path,BPC_pull,format_table,month_dic,year_dic
PL_path,Discrepancy_path,mapping_path,BPC_pull,format_table,month_dic,year_dic=Initial_Paramaters(operator)

@st.cache_resource
def Initial_Mapping(operator):
    # read account mapping
    account_mapping_obj =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
    account_mapping = pd.read_excel(account_mapping_obj['Body'].read(), sheet_name=sheet_name_account_mapping,header=0)   
    account_mapping["Tenant_Formated_Account"]=list(map(lambda x:x.upper().strip(),account_mapping["Tenant_Account"]))
    account_mapping=account_mapping[["Sabra_Account","Sabra_Second_Account","Tenant_Account","Tenant_Formated_Account","Conversion"]] 
    # read property mapping
    entity_mapping_obj =s3.get_object(Bucket=bucket_mapping, Key=mapping_path)
    entity_mapping=pd.read_excel(entity_mapping_obj['Body'].read(),sheet_name=sheet_name_entity_mapping,header=0)
    return entity_mapping,account_mapping
entity_mapping,account_mapping=Initial_Mapping(operator)

# Intialize a list of tuples containing the CSS styles for table headers
th_props = [('font-size', '14px'), ('text-align', 'left'),
            ('font-weight', 'bold'),('color', '#6d6d6d'),
            ('background-color', '#eeeeef'), ('border','1px solid #eeeeef')]

# Intialize a list of tuples containing the CSS styles for table data
td_props = [('font-size', '14px'), ('text-align', 'left')]

# Aggregate styles in a list
styles = [dict(selector="th", props=th_props),dict(selector="td", props=td_props)]

def left_align(s, props='text-align: left;'):
    return props
css='''
<style>
    section.main>div {
        padding-bottom: 1rem;
    }
    [data-testid="table"]>div>div>div>div>div {
        overflow: auto;
        height: 20vh;
    }
</style>
'''
st.markdown(css, unsafe_allow_html=True)

@st.cache_data
def Create_Tree_Hierarchy(bucket_mapping):
    #Create Tree select hierarchy
    parent_hierarchy_main=[{'label': "No need to map","value":"No need to map"}]
    parent_hierarchy_second=[{'label': "No need to map","value":"No need to map"}]
    BPCAccount = s3.get_object(Bucket=bucket_mapping, Key="Initial_info.xlsx")
    BPC_Account= pd.read_excel(BPCAccount['Body'].read(), sheet_name="BPC_Account_Info")
 
    for category in BPC_Account[BPC_Account["Type"]=="Main"]["Category"].unique():
        children_hierarchy=[]
        for account in BPC_Account[(BPC_Account["Category"]==category)&(BPC_Account["Type"]=="Main")]["Sabra_Account_Full_Name"]:
            dic={"label":account,"value":BPC_Account[(BPC_Account["Sabra_Account_Full_Name"]==account)&(BPC_Account["Type"]=="Main")]["BPC_Account_Name"].item()}
            children_hierarchy.append(dic)
        dic={"label":category, "value":category, "children":children_hierarchy}
        parent_hierarchy_main.append(dic)
        
    for category in BPC_Account[BPC_Account["Type"]=="Second"]["Category"].unique():
        children_hierarchy=[]
        for account in BPC_Account[(BPC_Account["Category"]==category)&(BPC_Account["Type"]=="Second")]["Sabra_Account_Full_Name"]:
            dic={"label":account,"value":BPC_Account.loc[(BPC_Account["Sabra_Account_Full_Name"]==account)&(BPC_Account["Type"]=="Second")]["BPC_Account_Name"].item()}
            children_hierarchy.append(dic)
        dic={"label":category,"value":category,"children":children_hierarchy}
        parent_hierarchy_second.append(dic)
    
    BPC_Account=BPC_Account[["BPC_Account_Name","Sabra_Account_Full_Name","Category"]]
    return parent_hierarchy_main,parent_hierarchy_second,BPC_Account
parent_hierarchy_main,parent_hierarchy_second,BPC_Account=Create_Tree_Hierarchy(bucket_mapping)

#-----------------------------------------------functions---------------------------------------------
# setting for page
@st.cache_data
def ChangeWidgetFontSize(wgt_txt, wch_font_size = '12px'):
    htmlstr = """<script>var elements = window.parent.document.querySelectorAll('*'), i;
                    for (i = 0; i < elements.length; ++i) { if (elements[i].innerText == |wgt_txt|) 
                        { elements[i].style.fontSize='""" + wch_font_size + """';} } </script>  """
    htmlstr = htmlstr.replace('|wgt_txt|', "'" + wgt_txt + "'")
    components.html(f"{htmlstr}", height=0, width=0)

@st.cache_data
def Identify_Tenant_Account_Col(PL,sheet_name,sheet_type):
    #search tenant account column in P&L, return col number of tenant account
    account_pool=account_mapping[["Sabra_Account","Tenant_Formated_Account"]].merge(BPC_Account[["BPC_Account_Name","Category"]], left_on="Sabra_Account", right_on="BPC_Account_Name",how="left")	       
    if sheet_type=="Sheet_Name":
        account_pool=account_pool.loc[account_pool["Sabra_Account"]!="NO NEED TO MAP"]["Tenant_Formated_Account"]
    elif sheet_type=="Sheet_Name_Occupancy": 
        account_pool=account_pool.loc[account_pool["Category"]=="Patient Days"]["Tenant_Formated_Account"]	       
    elif sheet_type=="Sheet_Name_Balance_Sheet":
        account_pool=account_pool.loc[account_pool["Category"]=="Balance Sheet"]["Tenant_Formated_Account"]
    
    for tenantAccount_col_no in range(0,PL.shape[1]):
        #trim and upper case 
        candidate_col=list(map(lambda x: str(x).strip().upper() if x==x else x,PL.iloc[:,tenantAccount_col_no]))
        #find out how many tenant accounts match with account_pool
        match=[x in candidate_col for x in account_pool]
        #If 10% of accounts match with account_mapping list, identify this col as tenant account.
        
        if len(match)>0 and sum(x for x in match)/len(match)>0.1:
            return tenantAccount_col_no  
        else:
            # it is the wrong account column, continue to check next column
            continue
            
    st.error("Fail to identify tenant accounts column in sheet—— '"+sheet_name+"'")
    st.stop()


def download_report(df,button_display):
    download_file=df.to_csv(index=False).encode('utf-8')
    st.download_button(label="Download "+button_display,data=download_file,file_name=operator+" "+button_display+".csv",mime="text/csv")
    
def Get_Year(single_string):
    if single_string!=single_string or single_string==None or type(single_string)==float:
        return 0,""
    else:
        for Year in year_dic.keys():
            for Year_keyword in year_dic[Year]:
                if Year_keyword in single_string:
                    return Year,Year_keyword
        return 0,""

def Get_Month_Year(single_string):
    if single_string!=single_string or single_string==None or type(single_string)==float:
        return 0,0
    if type(single_string)==datetime:
        return int(single_string.month),int(single_string.year)

    single_string=str(single_string).lower()
    Year,Year_keyword=Get_Year(single_string)
    
    # remove year from string, remove days from string
    single_string=single_string.replace(Year_keyword,"").replace("30","").replace("31","").replace("28","").replace("29","")
    
    for Month in month_dic.keys() :#[01,02,03...12]
        for  Month_keyword in month_dic[Month]: #['december','dec','12',...]
            if Month_keyword in single_string:
                remaining=single_string.replace(Month_keyword,"").replace("/","").replace("-","").replace(" ","").replace("_","")
                
                #if there are more than 3 other char in the string, this string is not month 
                if len(remaining)>=3:
                    return 0,0
                else:   
                    return Month,Year
            # string doesn't contain month keyword, continue to next month keyword
            else:
                continue
    # didn't find month. return month as 0
    return 0,Year    
    
def Month_continuity_check(month_list):
    inv=[]
    month_list=list(filter(lambda x:x!=0,month_list))
    month_len=len(month_list)
    if month_len==0:
        return False
    else:
        inv=[int(month_list[month_i+1])-int(month_list[month_i]) for month_i in range(month_len-1) ]
        #there are at most two types of difference in the month list which are in 1,-1,11,-11 
        if  len(set(inv))<=2 and all([x in [1,-1,11,-11] for x in set(inv)]):
            return True  # Month list is continous 
        else:
            return False # Month list is not continous 
            
def Year_continuity_check(year_list):
    inv=[]
    year_list=list(filter(lambda x:x!=0,year_list))
    year_len=len(year_list)
    if year_len==0:
        return False
    else:
        inv=[int(year_list[year_i+1])-int(year_list[year_i]) for year_i in range(year_len-1)]
        if len(set(inv))<=2 and all([x in [1,0,-1] for x in set(inv)]):
            return True         #years are continous
        else:
            return False

# add year to month_header: identify current year/last year giving a list of month
def Add_year_to_header(month_list):
    available_month=list(filter(lambda x:x!=0,month_list))
    today=date.today()
    current_year= today.year
    last_year=today.year-1
    if len(available_month)==1:
        
        if datetime.strptime(available_month[0]+"/01/"+current_year,'%m/%d/%Y').date()<today:
            year=current_year
        else:
            year=today.year-1
        return year
     
    year_change=0     
    # month decending  #and available_month[0]<today.month
    if (available_month[0]>available_month[1] and available_month[0]!=12) or (available_month[0]==1 and available_month[1]==12) : 
        date_of_assumption=datetime.strptime(str(available_month[0])+"/01/"+str(current_year),'%m/%d/%Y').date()
        if date_of_assumption<today and date_of_assumption.month<today.month:
            report_year_start=current_year
        elif date_of_assumption>=today:
            report_year_start=last_year
        for i in range(len(available_month)):
            available_month[i]=report_year_start-year_change
            if i<len(available_month)-1 and available_month[i+1]==12:
                year_change+=1
            
    # month ascending
    elif (available_month[0]<available_month[1] and available_month[0]!=12) or (available_month[0]==12 and available_month[1]==1): #and int(available_month[-1])<today.month
        date_of_assumption=datetime.strptime(str(available_month[-1])+"/01/"+str(current_year),'%m/%d/%Y').date()    
        if date_of_assumption<today:
            report_year_start=current_year
        elif date_of_assumption>=today:
            report_year_start=last_year
        for i in range(-1,len(available_month)*(-1)-1,-1):
   
            available_month[i]=report_year_start-year_change
            if i>len(available_month)*(-1) and available_month[i-1]==12:
                #print("year_change",year_change)
                year_change+=1
    else:
        return False
 
    j=0
    for i in range(len(month_list)):
        if month_list[i]!=0:
            month_list[i]=available_month[j]
            j+=1
    return month_list  

# search for the Month/year row and return row number

@st.cache_data
def Identify_Month_Row(PL,tenantAccount_col_no,sheet_name):
    PL_row_size=PL.shape[0]
    PL_col_size=PL.shape[1]
    search_row_size=min(15,PL_row_size)
    month_table=pd.DataFrame(0,index=range(search_row_size), columns=range(PL_col_size))
    year_table=pd.DataFrame(0,index=range(search_row_size), columns=range(PL_col_size))
    
    for row_i in range(search_row_size):
        for col_i in range(PL_col_size):
            month_table.iloc[row_i,col_i],year_table.iloc[row_i,col_i]=Get_Month_Year(PL.iloc[row_i,col_i])
    year_count=[]        
    month_count=[]
    max_len=0
    for row_i in range(search_row_size):
        # save the number of valid months of each row to month_count
        valid_month=list(filter(lambda x:x!=0,month_table.iloc[row_i,]))
        valid_year=list(filter(lambda x:x!=0,year_table.iloc[row_i,]))
        month_count.append(len(valid_month))
        year_count.append(len(valid_year))
        
    # can't find month keyword in any rows
    if all(map(lambda x:x==0,month_count)):
        st.error("Can't identify month/year columns in sheet——'"+sheet_name+"'")   
        st.stop()
        
    month_sort_index = np.argsort(np.array(month_count))
    year_sort_index = np.argsort(np.array(year_count))
    for month_index_i in range(-1,-4,-1): # only check three of the most possible rows
        #month_sort_index[-1] is the index number of month_count in which has max month count
        #month_sort_index[i] is also the index/row number of PL
        if month_count[month_sort_index[month_index_i]]>1:
            month_row=list(month_table.iloc[month_sort_index[month_index_i],])
           
	    # if True, it is the correct month row
            if Month_continuity_check(month_row):
		    
                for year_index_i in range(-1,-4,-1):
                    year_row=list(year_table.iloc[year_sort_index[year_index_i],])
		     # if month and year are not in the same places in the columns, year_row is not the correct one
                    if not all([year_row[i]==month_row[i] if month_row[i]==0 else year_row[i]!=0 for i in range(len(month_row))]):
                        continue
                    # check validation of year
                    if Year_continuity_check(year_row) \
                        and year_count[year_sort_index[year_index_i]]==month_count[month_sort_index[month_index_i]]:
                       
                        PL_date_header=year_table.iloc[year_sort_index[year_index_i],].apply(lambda x:str(int(x)))+\
                        month_table.iloc[month_sort_index[month_index_i],].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                        
                        return PL_date_header,month_sort_index[month_index_i]
                    
                    # all the year rows are not valid, add year to month
                    else:
                        continue

		# all the year rows are not valid, add year to month
                year_table.iloc[year_sort_index[year_index_i],]=Add_year_to_header(list(month_table.iloc[month_sort_index[month_index_i],]))
                PL_date_header=year_table.iloc[year_sort_index[year_index_i],].apply(lambda x:str(int(x)))+\
                month_table.iloc[month_sort_index[month_index_i],].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                original=PL.iloc[month_sort_index[month_index_i],]
                
                d_str = ''
                for i in range(len(PL_date_header)):
                        if PL_date_header[i]==0 or PL_date_header[i]=="0":
                            continue
                        else:
                            date=str(PL_date_header[i][4:6])+"/"+str(PL_date_header[i][0:4])
                            d_str +=",  "+str(original[i])+" — "+ date
                
                st.warning("Warning: Fail to identify 'year' in the month header of sheet '"+sheet_name+"'. Filled year as:")
                st.markdown(d_str[1:])
                return PL_date_header,month_sort_index[month_index_i]
                        
            # month is not continuous, check next one
            else:
                continue
                
        # only one month in header:month and year must exist for one month header
        elif month_count[month_sort_index[month_index_i]]==1:
            # month and year must match 
            st.write("There is only one month in sheet——'"+sheet_name+"'")
            col_month=0
            #find the col number of month
            while(month_table.iloc[month_sort_index[month_index_i],col_month]==0):
                col_month+=1
                
                #if month_table.iloc[month_sort_index[index_i],col_month]!=1:
                #if there is no year in month, continue 
            if  year_table.iloc[month_sort_index[month_index_i],col_month]==0:
                continue
           
            count_num=0
            count_str=0
            for row_month in range(month_sort_index[month_index_i],PL.shape[0]):
                if PL.iloc[row_month,col_month]==None or pd.isna(PL.iloc[row_month,col_month]) or PL.iloc[row_month,col_month]=="":
                    continue
                elif type(PL.iloc[row_month,col_month])==float or type(PL.iloc[row_month,col_month])==int:
                    count_num+=1
                else:
                    count_str+=1
                # count_num/str is count of numous/character data under month
                # for a real month column, numous data is supposed to be more than character data
            if count_str>0 and count_num/count_str<0.8:
                continue
                
            else:
                PL_date_header=year_table.iloc[month_sort_index[month_index_i],].apply(lambda x:str(int(x)))+\
                        month_table.iloc[month_sort_index[month_index_i],].apply(lambda x:"" if x==0 else "0"+str(int(x)) if x<10 else str(int(x)))
                        
                return PL_date_header,month_sort_index[month_index_i]
    st.error("Can't identify date row in P&L for sheet: '"+sheet_name+"'")
    st.stop()

def Save_File_toS3(uploaded_file, bucket, key):  
    try:
        s3.upload_fileobj(uploaded_file, bucket, key)
        st.success('{} successfully Uploaded'.format(uploaded_file.name))
        return True
    except FileNotFoundError:
        st.error("File can't be uploaded.")
        return False   
    
def Update_Sheet_inS3(bucket,key,sheet_name,df,how="replace"):  
    if how=="append":
        discrepancy_file =s3.get_object(Bucket=bucket, Key=key)
        original_df=pd.read_excel(discrepancy_file['Body'].read(), sheet_name=sheet_name,header=0)
        # remove original discrepancy and comments
        original_df = original_df.drop(original_df[original_df['Operator'] == operator].index)
	# update to new discrepancy and comments
        df = pd.concat([original_df,df]).reset_index(drop=True)
    load_file =s3.get_object(Bucket=bucket, Key=key)
    workbook = load_workbook(BytesIO(load_file['Body'].read()))
    workbook.remove(workbook[sheet_name])
    new_worksheet = workbook.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        new_worksheet.append(r)
	    
    with NamedTemporaryFile() as tmp:
         workbook.save(tmp.name)
         data = BytesIO(tmp.read())
    s3.upload_fileobj(data,bucket,key)

#@st.cache_data(experimental_allow_widgets=True)
def Manage_Property_Mapping(operator):
    global entity_mapping
    #all the properties are supposed to be in entity_mapping. 
    entity_mapping_updation=pd.DataFrame(columns=["Property_Name","Sheet_Name","Sheet_Name_Occupancy","Sheet_Name_Balance_Sheet"])
    number_of_property=entity_mapping.shape[0]
    with st.form(key="Mapping Properties"):
        col1,col2,col3,col4=st.columns([4,3,3,3])
        with col1:
            st.write("Property")
        with col2:
            st.write("P&L Sheetname")    
        with col3: 
            st.write("Occupancy Sheetname")    
        with col4:
            st.write("Balance sheet Sheetname")  
        for i in range(entity_mapping.shape[0]):
            col1,col2,col3,col4=st.columns([4,3,3,3])
            with col1:
                st.write("")
                st.write(entity_mapping.loc[i,"Property_Name"])
            with col2:
                entity_mapping_updation.loc[i,"Sheet_Name"]=st.text_input("",placeholder =entity_mapping.loc[i,"Sheet_Name"],key="P&L"+str(i))    
            with col3: 
                entity_mapping_updation.loc[i,"Sheet_Name_Occupancy"]=st.text_input("",placeholder =entity_mapping.loc[i,"Sheet_Name_Occupancy"],key="Census"+str(i))     
            with col4:
                entity_mapping_updation.loc[i,"Sheet_Name_Balance_Sheet"]=st.text_input("",placeholder =entity_mapping.loc[i,"Sheet_Name_Balance_Sheet"],key="BS"+str(i)) 
            submitted = st.form_submit_button("Submit")
            
    if submitted:
        for i in range(entity_mapping.shape[0]):
            if entity_mapping_updation.loc[i,"Sheet_Name"]:
                entity_mapping.loc[i,"Sheet_Name"]=entity_mapping_updation.loc[i,"Sheet_Name"] 
            if entity_mapping_updation.loc[i,"Sheet_Name_Occupancy"]:
                entity_mapping.loc[i,"Sheet_Name_Occupancy"]=entity_mapping_updation.loc[i,"Sheet_Name_Occupancy"]
            if  entity_mapping_updation.loc[i,"Sheet_Name_Balance_Sheet"]:
                entity_mapping.loc[i,"Sheet_Name_Balance_Sheet"]=entity_mapping_updation.loc[i,"Sheet_Name_Balance_Sheet"] 
        
        download_report(entity_mapping[["Property_Name","Sheet_Name","Sheet_Name_Occupancy","Sheet_Name_Balance_Sheet"]],"{} properties mapping".format(operator))
        # update account_mapping in S3     
        Update_Sheet_inS3(bucket_mapping,mapping_path,sheet_name_entity_mapping,entity_mapping)    
        return entity_mapping

@st.cache_data(experimental_allow_widgets=True)
def Manage_Account_Mapping(new_tenant_account):
    with st.form(key=new_tenant_account):
        col1,col2=st.columns(2) 
        with col1:
            st.write("Sabra main account")
            Sabra_main_account=streamlit_tree_select.tree_select(parent_hierarchy_main,only_leaf_checkboxes=True,key=new_tenant_account) 
        with col2:
            st.write("Sabra second account")
            Sabra_second_account= streamlit_tree_select.tree_select(parent_hierarchy_second,only_leaf_checkboxes=True,key=new_tenant_account+"1")
        submitted = st.form_submit_button("Submit")  
    if submitted:
        if len(Sabra_main_account['checked'])==1:
            Sabra_main_account_value=Sabra_main_account['checked'][0].upper()          
        elif len(Sabra_main_account['checked'])>1:
            st.warning("Only one to one mapping is allowed.")
            st.stop()
        elif Sabra_main_account['checked']==[]:
            st.warning("Please select Sabra main account for '{}'".format(new_tenant_account[i]))
            st.stop()
            
        if Sabra_second_account['checked']==[]:
            Sabra_second_account_value=''
        elif len(Sabra_second_account['checked'])==1:
            Sabra_second_account_value=Sabra_second_account['checked'][0].upper()
        elif len(Sabra_second_account['checked'])>1:
            st.warning("Only one to one mapping is allowed.")
            st.stop()
    else:
        st.stop()
                
    if Sabra_main_account_value=="NO NEED TO MAP":
        st.success("{} was successfully saved to 'No need to map' list.".format(new_tenant_account))
    elif Sabra_main_account_value:
        st.success("Successfully mapped '{}' to '{}'".format(new_tenant_account,Sabra_main_account_value))
    return Sabra_main_account_value,Sabra_second_account_value     

@st.cache_data(experimental_allow_widgets=True)
def Sheet_Process(entity_i,sheet_type,sheet_name):
    global account_mapping
    # read data from uploaded file
    count=0
    while(True):
        try:
            PL = pd.read_excel(uploaded_file,sheet_name=sheet_name,header=None)
            break
        except:
	    # if there is no sheet name for sold property in P&L, continue to process next property
            if entity_mapping.loc[entity_i,"DATE_SOLD_PAYOFF"]==entity_mapping.loc[entity_i,"DATE_SOLD_PAYOFF"]:
                return  pd.DataFrame(),pd.DataFrame()  
        
            col1,col2=st.columns(2) 
            with col1: 
                if sheet_type=="Sheet_Name":  
                    st.warning("Please provide sheet name of P&L data for property {}. ".format(entity_mapping.loc[entity_i,"Property_Name"]))
                elif sheet_type=="Sheet_Name_Occupancy":
                    st.warning("Please provide sheet name of Occupancy data for property {}. ".format(entity_mapping.loc[entity_i,"Property_Name"]))
                elif sheet_type=="Sheet_Name_Balance_Sheet":
                    st.warning("Please provide sheet name of Balance Sheet data in for property {}. ".format(entity_mapping.loc[entity_i,"Property_Name"]))
		    
            if len(PL_sheet_list)>0:
                with st.form(key=str(count)):                
                    sheet_name=st.selectbox(entity_mapping.loc[entity_i,"Property_Name"],[""]+PL_sheet_list)
                    submitted = st.form_submit_button("Submit")
            else:
                with st.form(key=str(count)):     
                    sheet_name = st.text_input(entity_mapping.loc[entity_i,"Property_Name"])
                    submitted = st.form_submit_button("Submit")
            if submitted:   
                count+=1
                continue
            else:
                st.stop()
    if count>0:
        # update sheet name in entity_mapping
        entity_mapping.loc[entity_i,sheet_type]=sheet_name  
        # update account_mapping in S3     
        Update_Sheet_inS3(bucket_mapping,mapping_path,sheet_name_entity_mapping,entity_mapping)    
    
    # Start checking process
    st.write("********Start to check property—'"+property_name+"' in sheet '"+sheet_name+"'********" )  
    tenantAccount_col_no=Identify_Tenant_Account_Col(PL,sheet_name,sheet_type)
    if tenantAccount_col_no==None:
        st.error("Fail to identify tenant account column in sheet '{}'".format(sheet_name))
        st.stop()    
    date_header=Identify_Month_Row(PL,tenantAccount_col_no,sheet_name)
  
    if len(date_header[0])==1 and date_header[0]==[0]:
        st.error("Fail to identify month/year header in sheet '{}', please add it and re-upload.".format(sheet_name))
        st.stop()     
    PL.columns=date_header[0]

    #set tenant_account as index of PL
    PL=PL.set_index(PL.iloc[:,tenantAccount_col_no].values)
   
    #remove row above date row and remove column without date col name
    PL=PL.iloc[date_header[1]+1:,PL.columns!='0']
    
    #remove rows with nan tenant account
    nan_index=list(filter(lambda x:x=="nan" or x=="" or x==" " or x!=x ,PL.index))
    PL.drop(nan_index, inplace=True)
    #set index as str ,strip
    PL.index=map(lambda x:str(x).strip(),PL.index)
    PL=PL.applymap(lambda x: 0 if (x!=x) or (type(x)==str) or x==" " else x)
    # remove columns with all nan/0
    PL=PL.loc[:,(PL!= 0).any(axis=0)]
    # remove rows with all nan/0 value
    PL=PL.loc[(PL!= 0).any(axis=1),:]
    return PL

@st.cache_data
def Mapping_PL_Sabra(PL,entity):
    # remove no need to map from account_mapping
    main_account_mapping=account_mapping.loc[list(map(lambda x:x==x and x.upper()!='NO NEED TO MAP',account_mapping["Sabra_Account"])),:]

    #concat main accounts with second accounts
    second_account_mapping=account_mapping.loc[(account_mapping["Sabra_Second_Account"]==account_mapping["Sabra_Second_Account"])&(account_mapping["Sabra_Second_Account"]!="NO NEED TO MAP")][["Sabra_Second_Account","Tenant_Formated_Account","Tenant_Account","Conversion"]].\
                           rename(columns={"Sabra_Second_Account": "Sabra_Account"})
    
    PL.index.name="Tenant_Account"
    PL["Tenant_Formated_Account"]=list(map(lambda x:x.upper() if type(x)==str else x,PL.index))
 
    PL=pd.concat([PL.merge(second_account_mapping,on="Tenant_Formated_Account",how='right'),PL.merge(main_account_mapping[main_account_mapping["Sabra_Account"]==main_account_mapping["Sabra_Account"]]\
                                            [["Sabra_Account","Tenant_Formated_Account","Tenant_Account","Conversion"]],on="Tenant_Formated_Account",how='right')])

    PL=PL.reset_index(drop=True)
    month_cols=list(filter(lambda x:str(x[0:2])=="20",PL.columns))
    for i in range(len(PL.index)):
        conversion=PL.loc[i,"Conversion"]
        if conversion!=conversion:
            continue
        else:
            for month in month_cols:
                before_conversion=PL.loc[i,month]
               
                if before_conversion!=before_conversion:
                    continue
                elif conversion=="/monthdays":		
                    PL.loc[i,month]=before_conversion/monthrange(int(str(month)[0:4]), int(str(month)[4:6]))[1]
                elif conversion[0]=="*":
                    PL.loc[i,month]= before_conversion*float(conversion.split("*")[0])
    PL=PL.drop(["Tenant_Formated_Account","Conversion"], axis=1)
    
    PL_with_detail=copy.copy(PL)
    PL_with_detail["Entity"]=entity
    PL_with_detail=PL_with_detail.set_index(['Entity', 'Sabra_Account',"Tenant_Account"])
    PL=PL.set_index("Sabra_Account",drop=True)
    PL=PL.drop(["Tenant_Account"], axis=1)
    # group by Sabra_Account
    PL=PL.groupby(by=PL.index).sum().replace(0,None)
    PL.index=[[entity]*len(PL.index),list(PL.index)]
    return PL,PL_with_detail
    
@st.cache_data
def Compare_PL_Sabra(Total_PL,PL_with_detail):
    PL_with_detail=PL_with_detail.reset_index(drop=False)
    diff_BPC_PL=pd.DataFrame(columns=["TIME","ENTITY","Sabra_Account","Sabra","P&L","Diff"])
    diff_BPC_PL_detail=pd.DataFrame(columns=["Entity","Sabra_Account","Tenant_Account","Month","P&L Value","Diff","Sabra"])
    for entity in entity_mapping["ENTITY"]:
        for matrix in BPC_Account.loc[(BPC_Account["Category"]!="Balance Sheet")]["BPC_Account_Name"]: 
            for timeid in Total_PL.columns.sort_values()[-2:]: # only compare two months
                try:
                    BPC_value=int(BPC_pull.loc[entity,matrix][timeid+'00'])
                except:
                    BPC_value=0
                try:
                    PL_value=int(Total_PL.loc[entity,matrix][timeid])
                except:
                    PL_value=0
                if BPC_value==0 and PL_value==0:
                    continue
                diff=BPC_value-PL_value
                if abs(diff)>=0.05*max(abs(PL_value),abs(BPC_value)):
                    diff_single_record=pd.DataFrame({"TIME":timeid,"ENTITY":entity,"Sabra_Account":matrix,"Sabra":BPC_value,\
                                                     "P&L":PL_value,"Diff":diff},index=[0])
                    
                    diff_BPC_PL=pd.concat([diff_BPC_PL,diff_single_record],ignore_index=True)

                    diff_detail_records=PL_with_detail.loc[(PL_with_detail["Sabra_Account"]==matrix)&(PL_with_detail["Entity"]==entity)]\
			                [["Entity","Sabra_Account","Tenant_Account",timeid]].rename(columns={timeid:"P&L Value"})
                    diff_detail_records["Month"]=timeid
                    diff_detail_records["Sabra"]=BPC_value
                    diff_detail_records["Diff"]=diff
                   
                    #if there is no record in diff_detail_records, means there is no mapping
                    if diff_detail_records.shape[0]==0:
                        diff_detail_records=pd.DataFrame({"Entity":entity,"Sabra_Account":matrix,"Tenant_Account":"Miss mapping accounts","Month":timeid,"Sabra":BPC_value,"Diff":diff,"P&L Value":0},index=[0])   
                    diff_BPC_PL_detail=pd.concat([diff_BPC_PL_detail,diff_detail_records])

    return diff_BPC_PL,diff_BPC_PL_detail

@st.cache_data(experimental_allow_widgets=True)
def View_Summary():
    global Total_PL
    def highlight_total(df):
        return ['color: blue']*len(df) if df.Sabra_Account.startswith("Total - ")  else ''*len(df)
    months=map(lambda x:x[4:6]+"/"+x[0:4],Total_PL.columns)
    m_str = ''
    for month in months:
        m_str += ", " + month
    st.write("Reporting months detected in P&L : "+m_str[1:])   
    st.write("The latest reporting month is:  "+latest_month[4:6]+"/"+latest_month[0:4])
    Total_PL.index=Total_PL.index.set_names(["ENTITY", "Sabra_Account"]) 
    Total_PL=Total_PL.fillna(0)
    latest_month_data=Total_PL[latest_month].reset_index(drop=False)
    latest_month_data=latest_month_data.merge(BPC_Account, left_on="Sabra_Account", right_on="BPC_Account_Name",how="left")
    latest_month_data=latest_month_data.merge(entity_mapping[["ENTITY","Property_Name"]], on="ENTITY",how="left")
    latest_month_data = latest_month_data.pivot(index=["Sabra_Account_Full_Name","Category"], columns="Property_Name", values=latest_month)
    latest_month_data.reset_index(drop=False,inplace=True)
    latest_month_data.rename(columns={"Sabra_Account_Full_Name":"Sabra_Account"},inplace=True) 
    latest_month_data = (pd.concat([latest_month_data.groupby(by='Category',as_index=False).sum().\
                       assign(Sabra_Account="Total_Sabra"),latest_month_data]).\
                         sort_values(by='Category', kind='stable', ignore_index=True)[latest_month_data.columns])
    latest_month_data=latest_month_data[latest_month_data["Sabra_Account"]==latest_month_data["Sabra_Account"]]
    
    for i in range(latest_month_data.shape[0]):
        if latest_month_data.loc[i,"Sabra_Account"]=="Total_Sabra":
            latest_month_data.loc[i,"Sabra_Account"]="Total - "+latest_month_data.loc[i,'Category']
        else:
            latest_month_data.loc[i,"Sabra_Account"]="        "+latest_month_data.loc[i,"Sabra_Account"]
    
    latest_month_data["Total"] = latest_month_data.drop(["Sabra_Account","Category"],axis=1).sum(axis=1)
   
    st.markdown(latest_month_data.drop(["Category"],axis=1).style.set_table_styles(styles).apply(highlight_total,axis=1).applymap(left_align)
		.format(precision=0,thousands=",").hide(axis="index").to_html(),unsafe_allow_html=True)
    download_report(latest_month_data,"{} {}-{} Reporting".format(operator,latest_month[4:6],latest_month[0:4]))

# can't use cache
def View_Discrepancy(percent_discrepancy_accounts): 
    global diff_BPC_PL
    if diff_BPC_PL.shape[0]>0:
        st.error("{0:.1f}% P&L data doesn't tie to Sabra data.  Please leave comments for each discrepancy in below table.".format(percent_discrepancy_accounts*100))
        edited_diff_BPC_PL = st.data_editor(
	diff_BPC_PL,
	width = 1200,
	column_order=("Property_Name","TIME","Sabra_Account_Full_Name","Sabra","P&L","Diff","Type comments below"),
	hide_index=True,
	disabled=("Property_Name","TIME","Sabra_Account_Full_Name","Sabra","P&L","Diff"),
	column_config={
       		"Sabra_Account_Full_Name": "Sabra_Account",
       		 "Property_Name": "Property",
		 "TIME":"Month",
		"P&L":st.column_config.TextColumn(
			"Tenant P&L",help="Tenant P&L is aggregated by detail tenant accounts connected with 'Sabra Account'"),
        	"Diff": st.column_config.TextColumn(
            		"Diff",help="Diff = Sabra-TenantP&L"),
		"Sabra": st.column_config.TextColumn(
            		"Sabra",help="Sabra data for previous month"),
		 "Type comments below":st.column_config.TextColumn(
            		"Type comments below",
            		help="Please provide an explanation and solution on discrepancy, like: confirm the changed. overwrite Sabra data with new one...",
			disabled =False,
            		required =False)
		}) 

        col1,col2=st.columns([1,3]) 
        with col1:
            submit_com=st.button("Submit comments")
        if submit_com:
            with col2:  
                with st.empty():
                    st.markdown("✔️ :green[Comments uploaded]")
                    time.sleep(1)
                    st.write(" ")
                Update_Sheet_inS3(bucket_PL,Discrepancy_path,sheet_name_discrepancy,edited_diff_BPC_PL,"append") 
            with col1:                        
                download_report(edited_diff_BPC_PL[["Property_Name","TIME","Sabra_Account_Full_Name","Sabra","P&L","Diff","Type comments below"]],"Discrepancy review")
    else:
        st.success("All previous data in P&L ties with Sabra data")

@st.cache_data(experimental_allow_widgets=True)  
def View_Discrepancy_Detail():
    global diff_BPC_PL,diff_BPC_PL_detail,Total_PL_detail,Total_PL
    # Sabra detail accounts mapping table
    def color_coding(row):
    	return ['color: blue'] * len(row) if row.Tenant_Account == " Total" else ['color: black'] * len(row)
    if diff_BPC_PL.shape[0]>0:
        st.markdown("---")
        st.markdown("P&L—Sabra detail accounts mapping (for discrepancy data)") 
        diff_BPC_PL_detail = (pd.concat([diff_BPC_PL_detail.groupby(["Entity","Sabra_Account","Month","Sabra","Diff"], as_index=False).sum()
                      .assign(Tenant_Account=" Total"),diff_BPC_PL_detail]).sort_values(by=["Entity","Sabra_Account","Month","Sabra","Diff"], kind='stable', ignore_index=True)[diff_BPC_PL_detail.columns])
        diff_BPC_PL_detail=diff_BPC_PL_detail.merge(BPC_Account[["BPC_Account_Name","Sabra_Account_Full_Name"]],left_on="Sabra_Account", right_on="BPC_Account_Name",how="left")
        diff_BPC_PL_detail=diff_BPC_PL_detail.merge(entity_mapping[["ENTITY","Property_Name"]],left_on="Entity", right_on="ENTITY",how="left")
        diff_BPC_PL_detail=diff_BPC_PL_detail[["Property_Name","Month","Sabra_Account_Full_Name","Tenant_Account","Sabra","P&L Value","Diff"]].\
			rename(columns={"Property_Name":"Property","Sabra_Account_Full_Name":"Sabra Account"})
        diff_BPC_PL_detail_for_download=diff_BPC_PL_detail.copy()
        for i in range(diff_BPC_PL_detail.shape[0]):
            if  diff_BPC_PL_detail.loc[i,"Tenant_Account"]!=" Total":
                diff_BPC_PL_detail.loc[i,"Property"]=""
                diff_BPC_PL_detail.loc[i,"Month"]=""
                diff_BPC_PL_detail.loc[i,"Sabra Account"]=""
                diff_BPC_PL_detail.loc[i,"Sabra"]=""
                diff_BPC_PL_detail.loc[i,"Diff"]=""
                diff_BPC_PL_detail.loc[i,"Tenant_Account"]="—— "+diff_BPC_PL_detail.loc[i,"Tenant_Account"]
        
        st.markdown(
            """
        <style type="text/css" media="screen">
        div[role="dataframe"] ul {
            height:300px;
        }
        </style>
            """,
        unsafe_allow_html=True )
        st.markdown(diff_BPC_PL_detail.style.set_table_styles(styles).apply(color_coding, axis=1).applymap(left_align)
		.format(precision=0,thousands=",").hide(axis="index").to_html(),unsafe_allow_html=True)	
        download_report(diff_BPC_PL_detail_for_download,"P&L accounts mapping for discrepancy")
    download_report(Total_PL_detail,"Full P&L accounts mapping")
   

@st.cache_data(experimental_allow_widgets=True)        
def PL_Process_Main(entity_i,sheet_type):  
    global latest_month
    #local sheet_name
    sheet_name=str(entity_mapping.loc[entity_i,sheet_type])
    if True:
            PL=Sheet_Process(entity_i,sheet_type,sheet_name)
         
            # mapping new tenant accounts
            new_tenant_account_list=list(filter(lambda x:x.upper().strip() not in list(account_mapping["Tenant_Formated_Account"]),PL.index))
            
            if len(new_tenant_account_list)>0:
                st.warning("Please complete mapping for below P&L accounts:")
                for i in range(len(new_tenant_account_list)):
                    st.markdown("## Map **'{}'** to Sabra account".format(new_tenant_account_list[i])) 
                    Sabra_main_account_value,Sabra_second_account_value=Manage_Account_Mapping(new_tenant_account_list[i])
                    #insert new record to the bottom line of account_mapping
                    account_mapping.loc[len(account_mapping.index)]=[Sabra_main_account_value,Sabra_second_account_value,new_tenant_account_list[i],new_tenant_account_list[i].upper(),"N"]           
                Update_Sheet_inS3(bucket_mapping,mapping_path,sheet_name_account_mapping,account_mapping) 
            
            #if there are duplicated accounts in P&L, ask for confirming
            dup_tenant_account=set([x for x in PL.index if list(PL.index).count(x) > 1])
            if len(dup_tenant_account)>0:
                for dup in dup_tenant_account:
                    if dup.upper() not in list(account_mapping[account_mapping["Sabra_Account"]=="NO NEED TO MAP"]["Tenant_Formated_Account"]):
                        st.warning("Warning: There are more than one '{}' accounts in sheet '{}'. They will be summed up by default.".format(dup,sheet_name))

            PL,PL_with_detail=Mapping_PL_Sabra(PL,entity_mapping.loc[entity_i,"ENTITY"])
            
            max_month_cols=str(max(list(PL.columns)))
	    # check the latest reporting month
            if latest_month=="2023":	    
                latest_month=max_month_cols
                
                col1,col2,col3=st.columns([4,1,6])
                with col1:
                    st.write("The latest reporting month is: {}/{}. Is it true?".format(latest_month[4:6],latest_month[0:4])) 
                with col2:		
                    y=st.button("Yes")          
                with col3:
                    n=st.button("No")   

                if n:
                    st.error("Please check the month header in sheet '{}' and make sure the latest or biggest month in month header is the new reporting month.".format(sheet_name))  
                    st.stop()
                elif not y:
                    st.stop()
            elif latest_month!=max_month_cols:
                st.error("The latest month in sheet '{}' is not {}. Please fix it and re-upload.".format(sheet_name,latest_month))
                st.stop()
	    # check the start reporting month
    return latest_month,PL,PL_with_detail

@st.cache_data(experimental_allow_widgets=True)  
def Upload_Section(uploaded_file):
    global PL_sheet_list,latest_month,property_name
    if True:
        if uploaded_file.name[-5:]=='.xlsx':
            PL_sheet_list=load_workbook(uploaded_file).sheetnames
        else:
            PL_sheet_list=[]
        
        if format_table["P&L_in_separate_sheets"][0]=="Y":
            Total_PL=pd.DataFrame()
            Total_PL_detail=pd.DataFrame()
            for entity_i in range(len(entity_mapping["ENTITY"])):
                sheet_name=str(entity_mapping.loc[entity_i,"Sheet_Name"])
                sheet_name_occupancy=str(entity_mapping.loc[entity_i,"Sheet_Name_Occupancy"])
                sheet_name_balance=str(entity_mapping.loc[entity_i,"Sheet_Name_Balance_Sheet"])
                property_name=str(entity_mapping.loc[entity_i,"Property_Name"])
		    
                latest_month,PL,PL_with_detail=PL_Process_Main(entity_i,"Sheet_Name")
		
        
		 # check if census data existed
                if sheet_name_occupancy!='nan' and sheet_name_occupancy==sheet_name_occupancy and sheet_name_occupancy!="" and sheet_name_occupancy!=" "\
                    and sheet_name_occupancy!=sheet_name:
                    latest_month,PL_occ,PL_with_detail_occ=PL_Process_Main(entity_i,"Sheet_Name_Occupancy") 
                    PL=PL.combine_first(PL_occ)
                    PL_with_detail=PL_with_detail.combine_first(PL_with_detail_occ)
                    #PL_with_detail= PL_with_detail.loc[(PL_with_detail!= None).any(axis=1),:]
		# check if balance sheet data existed   
		
                if sheet_name_balance!='nan' and sheet_name_balance==sheet_name_balance and sheet_name_balance!="" and sheet_name_balance!=" " and sheet_name_balance!=sheet_name:
                        latest_month,PL_BS,PL_with_detail_BS=PL_Process_Main(entity_i,"Sheet_Name_Balance_Sheet")
                        PL=PL.combine_first(PL_BS)
                        # remove rows with all None value
                        #PL= PL.loc[(PL!= None).any(axis=1),:]
                        PL_with_detail=PL_with_detail.combine_first(PL_with_detail_BS)
                        #PL_with_detail= PL_with_detail.loc[(PL_with_detail!= None).any(axis=1),:]
                
                
                Total_PL=pd.concat([Total_PL,PL], ignore_index=False, sort=False)
                Total_PL_detail=pd.concat([Total_PL_detail,PL_with_detail], ignore_index=False, sort=False)
                st.success("Property {} checked.".format(entity_mapping.loc[entity_i,"Property_Name"]))

            # if Sheet_Name_Occupancy is available, process occupancy data separately
	    # check if census data existed
		
            diff_BPC_PL,diff_BPC_PL_detail=Compare_PL_Sabra(Total_PL,Total_PL_detail)
	    # save uploaded tenant file to S3
            Save_File_toS3(uploaded_file,bucket_PL,PL_path)
            
            if diff_BPC_PL.shape[0]>0:
                percent_discrepancy_accounts=diff_BPC_PL.shape[0]/(BPC_Account.shape[0]*len(Total_PL.columns))
                diff_BPC_PL=diff_BPC_PL.merge(BPC_Account,left_on="Sabra_Account",right_on="BPC_Account_Name",how="left")        
                diff_BPC_PL=diff_BPC_PL.merge(entity_mapping, on="ENTITY",how="left")
                diff_BPC_PL['Type comments below']=""
                diff_BPC_PL['Operator']=operator
    return Total_PL,Total_PL_detail,diff_BPC_PL,diff_BPC_PL_detail,percent_discrepancy_accounts
#----------------------------------website widges------------------------------------
menu=["Upload P&L","Manage Mapping","Instructions"]
choice=st.sidebar.selectbox("Menu", menu)
status_record=pd.DataFrame(columns=["Entity","BS","Revenue",""])

if choice=="Upload P&L" and operator!='select operator':
    st.subheader("Upload P&L:")
    col1,col2=st.columns(2) 
    with col1:
        with st.form("my-form", clear_on_submit=True):
            uploaded_file=st.file_uploader(":star: :red[XLSX recommended] :star:",type={"xlsx", "xlsm","xls"},accept_multiple_files=False)
            col3,col4=st.columns([1,3]) 
            with col3:
                submitted = st.form_submit_button("Upload")
            with col4:
                if submitted:
		# clear cache for every upload
                    st.cache_data.clear()
                    st.cache_resource.clear()
                    st.write("{} uploaded.".format(uploaded_file.name))
     
    if uploaded_file:
	# initial parameter
        TENANT_ID=format_table["Tenant_ID"][0]
        global latest_month
        latest_month="2023"
        Total_PL,Total_PL_detail,diff_BPC_PL,diff_BPC_PL_detail,percent_discrepancy_accounts=Upload_Section(uploaded_file)

        # 1 Summary
        with st.expander("Summary of P&L" ,expanded=True):
            ChangeWidgetFontSize('Summary of P&L', '25px')
            View_Summary()
        
        # 2 Discrepancy of Historic Data
        with st.expander("Discrepancy for Historic Data",expanded=True):
            ChangeWidgetFontSize('Discrepancy for Historic Data', '25px')
            View_Discrepancy(percent_discrepancy_accounts)
            View_Discrepancy_Detail()
    time.sleep(200)               
	
elif choice=="Manage Mapping" and operator!='select operator':
    with st.expander("Manage Property Mapping" ,expanded=True):
        ChangeWidgetFontSize('Manage Property Mapping', '25px')
        entity_mapping=Manage_Property_Mapping(operator)
    with st.expander("Manage Account Mapping",expanded=True):
        ChangeWidgetFontSize('Manage Account Mapping', '25px')
        col1,col2=st.columns(2)
        with col1:
            new_tenant_account=st.text_input("Enter new tenant account and press enter to apply:")
        if new_tenant_account:
            st.markdown("## Map **'{}'** to Sabra account".format(new_tenant_account)) 
            Sabra_main_account_value,Sabra_second_account_value=Manage_Account_Mapping(new_tenant_account)
            #insert new record to the bottom line of account_mapping
            account_mapping.loc[len(account_mapping.index)]=[Sabra_main_account_value,Sabra_second_account_value,new_tenant_account,new_tenant_account.upper(),"N"]   
            Update_Sheet_inS3(bucket_mapping,mapping_path,sheet_name_account_mapping,account_mapping)
time.sleep(5000) 
